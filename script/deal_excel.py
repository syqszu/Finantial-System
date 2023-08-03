import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment
import warnings
import os

warnings.filterwarnings('ignore')

global excel_in_url, excel_out_url, union_outdir, union_not_outdir, union_yet_outdir, union_tofill_outdir, \
    match_out_col, match_in_col, temp_col, union_yet_df, union_not_df

# 配置进项文件、销项文件还有输出文件的路径。
excel_in_url = r"D:\dev\ExcelUtil\temp\3月份进项.xlsx"
excel_out_url = r"D:\dev\ExcelUtil\temp\3月份销项.xlsx"
union_outdir = r"D:\dev\ExcelUtil\temp\out"
temp_col = pd.DataFrame()







# -*- coding: UTF-8 -*-
# 具体匹配过程，返回月匹配完成表和月匹配剩余表
# union_yet_outdir, union_not_outdir 分别是月匹配完成表和剩余表的路径。
def get_col():
    # 月匹配完成表
    global union_yet_df, union_not_df
    union_yet_df = pd.DataFrame(
        columns=['开票日期', '序号', '发票号码', '税收分类编码', '货物、应税劳务及服务', '规格型号', '单位',
                 '上月原数量', '上月程序出库数', '上月人工出库数', '本月剩余数量', '不含税单价',
                 '不含税金额', '含税单价', '含税总金额', '供应商',  # 16 销项

                 '开票日期1', '序号1', '发票日期1', '税收分类编码1', '货物、应税劳务及服务1', '规格型号1', '单位1',
                 '上月原数量1', '上月程序出库数1', '上月人工出库数1', '本月剩余数量1', '不含税单价1',
                 '不含税金额1', '含税单价1', '含税总金额1', '销方名称'  # 进项
                 ])
    union_not_df = union_yet_df.copy()  # 月剩余待匹配表
    id = 1
    # 暴力循环
    print("月匹配完成表开始生成")
    for idx_out, row_out in match_out_col.iterrows():
        for idx_in, row_in in match_in_col.iterrows():

            date_out = row_out["开票日期"]  # 2023-03-05
            date_in = row_in["开票日期"]
            number_out = row_out["发票号码"]
            number_in = row_in["发票号码"]
            loads_out = row_out["货物、应税劳务及服务"]
            loads_in = row_out["货物、应税劳务及服务"]
            size_out = row_out["规格型号"]
            size_in = row_out["规格型号"]
            unit_out = row_out["单位"]
            unit_in = row_out["单位"]
            origin_nums_out = row_out["数量"]  # 销项上月原数量
            origin_nums_in = row_in["数量"]  # 进项上月原数量

            notax_perprice_out = float(row_out["单价"])  # 销项不含税单价
            notax_perprice_in = float(row_in["单价"])  # 进项不含税单价

            notax_money_out = row_out["金额"]  # 销项不含税金额
            notax_money_in = row_in["金额"]  # 进项不含税金额

            tax_out = float(row_out["税率"].strip("%")) / 100  # 销项税率
            tax_in = float(row_in["税率"].strip("%")) / 100  # 进项税率

            # print(type(row_in["税率"]))
            # print(row_in["税率"])

            tax_perprice_out = notax_perprice_out * (1 + tax_out)  # 销项含税单价
            tax_perprice_in = notax_perprice_in * (1 + tax_in)  # 进项含税单价

            tax_money_out = row_out["价税合计"]  # 销项含税总金额
            tax_money_in = row_in["价税合计"]  # 进项含税总金额

            name_out = row_in["销方名称"]  # 销方名称
            name_in = row_out["购方名称"]  # 供应商

            difference = abs(tax_perprice_in - tax_perprice_out)
            percentage_differ = (difference / tax_perprice_in) * 100

            case1 = row_out[0] == row_in[0] and row_out["vis"] == 0 and row_in["vis"] == 0

            # front7_in = "{:.7f}".format(row_in[0])
            front7_in = row_in[0][0:6]
            # print(front7_in)
            # print(type(front7_in))

            # front7_out = "{:.7f}".format(row_out[0])
            front7_out = row_out[0][0:6]

            case2 = front7_in == front7_out and row_out["vis"] == 0 and row_in["vis"] == 0 and percentage_differ <= 10

            # 第一次遍历，将税收分类编码完美匹配的vis打上1的标记   前七位相同，含税单价相差不超过10%的vis打上2的标记
            if case1 or case2:

                yet_row = {"开票日期": date_out, "序号": id, "发票号码": number_out,
                           "税收分类编码": row_out[0], "货物、应税劳务及服务": loads_out, "规格型号": size_out,
                           "单位": unit_out, "上月原数量": origin_nums_out}

                # value = [date_out, number_out, loads_out, size_out, unit_out, origin_nums_out]
                # union_yet_df = pd.DataFrame(np.insert(union_yet_df.values, id-1, values=value), axis=0)

                outbound_pro_nums_out = outbound_pro_nums_in = 0  # 上月程序出库数
                remain_nums_out = remain_nums_in = 0  # 本月剩余数量

                if origin_nums_in < origin_nums_out:  # 进项上月原数量<销项上月原数量
                    outbound_pro_nums_out = outbound_pro_nums_in = origin_nums_in  # 销进项上月程序出库数=进项上月原数量
                    remain_nums_out = origin_nums_out - origin_nums_in  # 销项本月剩余量=销项上月原数量-进项上月原数量
                    remain_nums_in = 0  # 进项本月剩余量=0
                    temp_row = {}
                    for col in out_col:
                        temp_row.update({col: row_out[col]})
                    temp_row.update({"数量": remain_nums_out})
                    match_out_col.loc[len(match_out_col)] = temp_row

                else:
                    outbound_pro_nums_out = outbound_pro_nums_in = origin_nums_out
                    remain_nums_in = origin_nums_in - origin_nums_out
                    remain_nums_out = 0
                    temp_row = {}
                    for col in in_col:
                        temp_row.update({col: row_in[col]})
                    temp_row.update({"数量": remain_nums_in})
                    match_in_col.loc[len(match_in_col)] = temp_row

                yet_row.update(
                    {"上月程序出库数": outbound_pro_nums_out, "上月人工出库数": 0, "本月剩余数量": remain_nums_out})

                yet_row.update(
                    {"不含税单价": notax_perprice_out, "不含税金额": notax_money_out, "含税单价": tax_perprice_out,
                     "含税总金额": tax_money_out, "销方名称": name_out})
                yet_row.update({"开票日期1": date_in, "序号1": id, "发票号码1": number_in, "税收分类编码1": row_in[0],
                                "货物、应税劳务及服务1": loads_in, "规格型号1": size_in, "单位1": unit_in,
                                "上月原数量1": origin_nums_in, "上月程序出库数1": outbound_pro_nums_in,
                                "上月人工出库数1": 0,
                                "本月剩余数量1": remain_nums_in, "不含税单价1": notax_perprice_in,
                                "不含税金额1": notax_money_in,
                                "含税单价1": tax_perprice_in, "含税金额1": tax_money_in, "供应商": name_in})

                union_yet_df = union_yet_df.append(yet_row, ignore_index=True)

                if case1:
                    match_out_col.at[idx_out, "vis"] = match_in_col.at[idx_in, "vis"] = 1
                if case2:
                    match_out_col.at[idx_out, "vis"] = match_in_col.at[idx_in, "vis"] = 2

                id += 1
    print("月匹配完成表生成完毕")
    print("月匹配剩余表开始生成")
    # 生成未匹配表
    id = 1
    for idx_out, row_out in match_out_col.iterrows():
        if row_out["vis"] == 0:
            date_out = row_out["开票日期"]  # 2023-03-05
            number_out = row_out["发票号码"]
            loads_out = row_out["货物、应税劳务及服务"]
            size_out = row_out["规格型号"]
            unit_out = row_out["单位"]
            origin_nums_out = remain_nums_out = row_out["数量"]  # 销项上月原数量=本月剩余数量=销项文件的数量
            notax_perprice_out = float(row_out["单价"])  # 销项不含税单价
            notax_money_out = row_out["金额"]  # 销项不含税金额
            tax_out = float(row_out["税率"].strip("%")) / 100  # 销项税率
            tax_perprice_out = notax_perprice_out * (1 + tax_out)  # 销项含税单价
            tax_money_out = row_out["价税合计"]  # 销项含税总金额
            name_out = row_out["购方名称"]  # 供应商
            not_row = {"开票日期": date_out, "序号": id, "发票号码": number_out,
                       "税收分类编码": row_out[0], "货物、应税劳务及服务": loads_out, "规格型号": size_out,
                       "单位": unit_out, "上月原数量": origin_nums_out, "上月程序出库数": 0,
                       "上月人工出库数": 0, "本月剩余数量": remain_nums_out, "不含税单价": notax_perprice_out,
                       "不含税金额": notax_money_out, "含税单价": tax_perprice_out,
                       "含税总金额": tax_money_out, "供应商": name_out}
            union_not_df = union_not_df.append(not_row, ignore_index=True)
            id += 1
        # union_not_df = union_not_df.append(not_row, ignore_index=True)

    id = 1
    for idx_in, row_in in match_in_col.iterrows():
        if row_in["vis"] == 0:
            date_in = row_in["开票日期"]
            number_in = row_in["发票号码"]
            loads_in = row_out["货物、应税劳务及服务"]
            size_in = row_out["规格型号"]
            unit_in = row_out["单位"]
            origin_nums_in = remain_nums_in = row_in["数量"]  # 进项上月原数量=本月剩余数量=销项文件的数量
            notax_perprice_in = float(row_in["单价"])  # 进项不含税单价
            notax_money_in = row_in["金额"]  # 进项不含税金额
            tax_in = float(row_in["税率"].strip("%")) / 100  # 进项税率
            tax_perprice_in = notax_perprice_in * (1 + tax_in)  # 进项含税单价
            tax_money_in = row_in["价税合计"]  # 进项含税总金额
            name_in = row_in["销方名称"]  # 销方名称
            not_row = {"开票日期1": date_in, "序号1": id, "发票号码1": number_in, "税收分类编码1": row_in[0],
                       "货物、应税劳务及服务1": loads_in, "规格型号1": size_in, "单位1": unit_in,
                       "上月原数量1": origin_nums_in, "上月程序出库数1": 0,
                       "上月人工出库数1": 0,
                       "本月剩余数量1": remain_nums_in, "不含税单价1": notax_perprice_in,
                       "不含税金额1": notax_money_in,
                       "含税单价1": tax_perprice_in, "含税总金额1": tax_money_in, "销方名称": name_in}
            union_not_df = union_not_df.append(not_row, ignore_index=True)
    # print(union_yet_df)
    print("月匹配剩余表生成完毕")
    # 1080207990000000000 是完全一致的
    # union_yet_df.to_excel(union_yet_outdir, index=False)
    # union_not_df.to_excel(union_not_outdir, index=False)


def create_union_excel():
    # 创建一个新的工作簿 选择活动工作表
    wb = openpyxl.Workbook()
    ws = wb.active

    # 定义填充颜色
    blue_fill = PatternFill(fill_type='solid', fgColor='FFDDEBF7')
    orange_fill = PatternFill(fill_type='solid', fgColor='FFFFE699')

    # 填充A-P列为蓝色 Q-AF列为橙色
    for col in range(1, 17):
        ws.cell(row=1, column=col).fill = blue_fill
    for col in range(17, 33):
        ws.cell(row=1, column=col).fill = orange_fill

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    ws.cell(row=1, column=1).value = "销项"
    ws.cell(row=1, column=1).fill = blue_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # 合并Q-AF列，并设置文本为"进项"
    ws.merge_cells(start_row=1, start_column=17, end_row=1, end_column=32)
    ws.cell(row=1, column=17).value = "进项"
    ws.cell(row=1, column=17).fill = orange_fill
    ws.cell(row=1, column=17).alignment = Alignment(horizontal='center')

    # 定义列名列表
    column_names = ["开票日期", "序号", "发票号码", "税收分类编码", "货物、应税劳务及服务", "规格型号", "单位",
                    "上月原数量", "上月程序出库数",
                    "上月人工出库数", "本月剩余数量", "不含税单价", "不含税金额", "含税单价", "含税总金额", "销方名称",
                    "开票日期", "序号", "发票号码", "税收分类编码", "货物、应税劳务及服务", "规格型号", "单位",
                    "上月原数量", "上月程序出库数",
                    "上月人工出库数", "本月剩余数量", "不含税单价", "不含税金额", "含税单价", "含税总金额", "供应商"]
    # 设置第二行的列名
    for index, name in enumerate(column_names):
        ws.cell(row=2, column=index + 1).value = name

    bigger_columns = ['D', 'H', 'I', 'J', 'K', 'L', 'M', 'O', 'T', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AE']
    for column in bigger_columns:
        ws.column_dimensions[column].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['AF'].width = 25

    # 保存文件
    wb.save('union_table.xlsx')


if __name__ == '__main__':
    print("开始配置输出路径")
    getOutdir()
    print("得到输出路径")
    print("开始预处理")
    pre_deal()
    print("预处理完毕")
    get_col()
    # getOutdir()
    # create_union_excel()
