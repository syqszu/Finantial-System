# _*_ coding : utf-8 _*_
# @Time : 2023/4/24 16:15
# @Author : Origami
# @File : test
# @Project : ExcelUtil
import time

import pandas as pd
import numpy as np
import openpyxl as opx
# from ReadYaml import get_yaml_data
from .ReadYaml import get_yaml_data
from datetime import datetime

# in_excel，out_excel，union_excel分别表示进项表、销项表、合并后的表格
def solution(in_excel, out_excel, union_excel):
    # 给文件备份
    # backup_file(in_excel, out_excel, union_excel)
    # in_col 和 out_col 需要从配置文件中读取
    # match_in: 4/5/6 货物、应税劳务及服务/规格型号/数量
    # match_out: 5/6/7 货物、应税劳务及服务/规格型号/数量
    match_out, match_in = pre_process(in_excel, out_excel, union_excel)
    idx = 0
    # 存放匹配成功
    res = []
    while idx < len(match_out):
        if idx != 0 and match_out[idx][4] == match_out[idx - 1][4] and match_out[idx][5] == match_out[idx - 1][5]:
            # 若待匹配的销项和前一个销项数据的货物名字类型均相同，说明可以跳过
            idx += 1
            continue
        # 匹配操作，从match_in[0] ~ match[len(match_in) - 1]依次match_out[0 ~ len(match_out) - 1]

        # 从match_in列表依次匹配当前销项数据。
        out_goods_name = match_out[idx][4]
        out_goods_type = match_out[idx][5]
        match_success = False
        for i in range(len(match_in)):
            if match_success:
                i = 0
            if i > len(match_in):
                break
            in_goods_name = match_in[i][5]
            in_goods_type = match_in[i][6]
            if match(out_goods_name, out_goods_type, in_goods_name, in_goods_type):
                match_success = True
                out_goods_num = float(match_out[idx][6])
                in_goods_num = float(match_in[i][7])
                if out_goods_num > in_goods_num:
                    # 拷贝一份
                    match_out_copy = copy_arr(match_out[idx])
                    # 如果销项更多，那么删掉进项， 减去部分销项

                    # 将float64转成float
                    match_out_copy[6] = np.float(match_out_copy[6])
                    match_out[idx][6] = np.float(match_out[idx][6])

                    # 1.拷贝 2.减去匹配的
                    match_out_copy[6] = in_goods_num
                    match_out[idx][6] -= in_goods_num
                    res.append(combine_arr(match_out_copy, match_in[i], False, 0, 0))
                    match_in = np.delete(np.array(match_in, dtype=object), i, axis=0).tolist()
                elif out_goods_num < in_goods_num:
                    # 如果进项更多，那么删掉销项， 减去部分进项
                    match_in_copy = copy_arr(match_in[i])

                    match_in_copy[7] = np.float(match_in_copy[7])
                    match_in[i][7] = np.float(match_in[i][7])

                    match_in_copy[7] = out_goods_num
                    match_in[i][7] -= out_goods_num
                    res.append(combine_arr(match_out[idx], match_in_copy, False, 0, 0))
                    match_out = np.delete(np.array(match_out, dtype=object), idx, axis=0).tolist()
                else:
                    # 一样多，都删掉
                    res.append(combine_arr(match_out[idx], match_in[i], False, 0, 0))
                    match_in = np.delete(np.array(match_in, dtype=object), i, axis=0).tolist()
                    match_out = np.delete(np.array(match_out, dtype=object), idx, axis=0).tolist()
        if not match_success:
            idx += 1

    # 计算res表的进销对比区
    print(res)

    output_excel(res, match_out, match_in)


def output_excel(res, match_out, match_in):
    res_excel_name = '进、销项完成匹配表.xlsx'
    merge_excel_name = '总（待）匹配表.xlsx'
    # 获取文件输出路径
    out_path = get_yaml_data('config.yaml')['out_path']
    # 获取df
    comparison_df = get_comparison(res)
    merged_df = get_merge(match_out, match_in)
    res = get_res(res)
    # 合并匹配结果和对比结果
    res = pd.concat([res, comparison_df], axis=1, ignore_index=True)
    # 涂颜色
    res = res.style.apply(
        lambda x: ['background-color: #FFF2CC' if i <= 10 else (
            'background-color: #D9E1F2' if i <= 22 else 'background-color: #E2EFDA') for i in range(len(x))],
        axis=1)
    merged_df = merged_df.style.apply(
        lambda x: ['background-color: #FFF2CC' if i <= 10 else 'background-color: #E2EFDA' for i in range(len(x))],
        axis=1
    )
    # 写入文件
    res.to_excel(out_path + res_excel_name, index=False, header=False)
    merged_df.to_excel(out_path + merge_excel_name, index=False, header=False)
    # 画边框
    draw_border(out_path + res_excel_name)
    draw_border(out_path + merge_excel_name)
    # 合并居中单元格
    combine_and_center_cell(out_path + res_excel_name, 'A1', 'K1')
    combine_and_center_cell(out_path + res_excel_name, 'L1', 'W1')
    combine_and_center_cell(out_path + res_excel_name, 'X1', 'AA1')

    combine_and_center_cell(out_path + merge_excel_name, 'A1', 'K1')
    combine_and_center_cell(out_path + merge_excel_name, 'L1', 'W1')

    # merged_df.to_excel(out_path + "merge.xlsx", index=False, header=my_header)


def combine_and_center_cell(file_name, left, right):
    wb = opx.load_workbook(file_name)
    # 获取第一个工作表
    ws = wb.active
    # 合并单元格
    ws.merge_cells(left + ':' + right)
    # 居中合并的单元格
    ws[left].alignment = opx.styles.Alignment(horizontal='center', vertical='center')
    # 保存文件
    wb.save(file_name)


def draw_border(file_name):
    df = pd.read_excel(
        file_name)  # 打开Excel文件
    wb = opx.load_workbook(file_name)  # 选择工作表
    ws = wb.active  # 画框
    border = opx.styles.borders.Border(left=opx.styles.borders.Side(style='thin'),
                                       right=opx.styles.borders.Side(style='thin'),
                                       top=opx.styles.borders.Side(style='thin'),
                                       bottom=opx.styles.borders.Side(style='thin'))
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
    # 保存Excel文件
    wb.save(file_name)


def get_merge(match_out, match_in):
    empty_in = ['', '', '', '', '', '', '', '', '', '', '', '']
    empty_out = ['', '', '', '', '', '', '', '', '', '', '']
    match_in.append(empty_in)
    match_out.append(empty_out)
    match_out = pd.DataFrame(match_out)
    match_in = pd.DataFrame(match_in)
    merged_df = pd.concat([match_out, match_in], axis=1, ignore_index=True)  # 横向合并两个excel表
    res_head = [
        ['进项', '进项', '进项', '进项', '进项', '进项', '进项', '进项', '进项', '进项', '进项', '销项', '销项', '销项',
         '销项', '销项', '销项', '销项', '销项', '销项', '销项', '销项', '销项'],
        ['月份', '序号', '发票号码', '开票日期', '货物、应税劳务及服务', '规格型号', '数量', '不含税单价',
         '单位', '不含税金额', '含税总金额', '月份', '序号', '发票号码', '开票日期', '销方名称',
         '货物、应税劳务及服务', '规格型号', '数量', '不含税单价', '单位', '不含税金额', '含税金额']]
    res_head = pd.DataFrame(res_head)
    merged_df = pd.concat([res_head, merged_df], axis=0, ignore_index=True)  # 纵向合并
    return merged_df


def get_res(res):
    res_head = [
        ['进项', '进项', '进项', '进项', '进项', '进项', '进项', '进项', '进项', '进项', '进项', '销项', '销项', '销项',
         '销项', '销项', '销项', '销项', '销项', '销项', '销项', '销项', '销项'],
        ['月份', '序号', '发票号码', '开票日期', '货物、应税劳务及服务', '规格型号', '数量', '不含税单价',
         '单位', '不含税金额', '含税总金额', '月份', '序号', '发票号码', '开票日期', '销方名称',
         '货物、应税劳务及服务', '规格型号', '数量', '不含税单价', '单位', '不含税金额', '含税金额']]
    res = pd.DataFrame(res)
    res_head = pd.DataFrame(res_head)
    res = pd.concat([res_head, res], axis=0, ignore_index=True)  # 纵向合并
    return res


def get_comparison(arr):
    comparison = [['进销比对区', '进销比对区', '进销比对区', '进销比对区'], [
        '进项含税单价', '销项含税单价', '进项含税总金额', '销项含税金额']]
    for a in arr:
        single_price_in = float(a[22]) / float(a[18])
        single_price_out = float(a[10]) / float(a[6])
        sum_price_in = a[22]
        sum_price_out = a[10]
        comparison.append([single_price_in, single_price_out, sum_price_in, sum_price_out])

    print(comparison)
    df = pd.DataFrame(comparison)
    df = df.style.apply(
        lambda x: [
            'background-color: #E2EFDA; color: black; font-weight: bold' if i == 0 else 'background-color: #E2EFDA; color: black; '
            for i in range(len(x))
        ], axis=0
    )
    return df.data


def match(name1, type1, name2, type2):
    return name1 == name2 and type1 == type2


def pre_process(in_excel, out_excel, union_excel):
    # 首先调用get_col函数获取需要读取的列索引。
    in_col, out_col, match_in_col, match_out_col = get_col()

    # 获取当月月份
    cur_month = str(datetime.today().month) + "月"
    print(cur_month)
    # 读取第一行开始的指定列。
    df1 = pd.read_excel(in_excel, header=1, usecols=in_col)  # 读取第一个Excel文件
    df2 = pd.read_excel(out_excel, header=1, usecols=out_col)  # 读取第二个Excel文件
    df3_in = pd.read_excel(union_excel, header=1, usecols=match_in_col)
    df3_out = pd.read_excel(union_excel, header=1, usecols=match_out_col)
    # 将读取到的数据转换为数组
    arr1 = get_arr(df1)
    arr2 = get_arr(df2)
    arr3_in = get_arr(df3_in)
    arr3_out = get_arr(df3_out)
    # 合并
    match_out = combine_arr(arr3_out, arr2, True, cur_month, len(arr3_out) + 1)
    match_in = combine_arr(arr3_in, arr1, True, cur_month, len(arr3_in) + 1)

    return match_out, match_in


def get_col():
    out_col = ["发票号码", "开票日期", "货物、应税劳务及服务", "规格型号", "数量", "单价", "单位", "金额",
               "价税合计"]
    in_col = ["发票号码", "开票日期", "销方名称", "货物、应税劳务及服务", "规格型号", "数量", "单价", "单位", "金额",
              "价税合计"]

    match_out_col_num = 11
    match_out_col = []
    for i in range(match_out_col_num):
        match_out_col.append(i)
    match_in_col_num = 12
    match_in_col = []
    for i in range(11, 11 + match_in_col_num):
        match_in_col.append(i)
    return in_col, out_col, match_in_col, match_out_col


def copy_arr(arr):
    res = []
    for a in arr:
        res.append(a)
    return res


def get_arr(df):
    arr = []
    for i in range(len(df)):
        # 跳过nan
        arr.append(df.loc[i].to_list())
    for i in range(len(df) - 1, -1, -1):
        if pd.isnull(arr[i][0]):
            arr.pop()
    return arr


def get_index(dfs, field1, field2):
    idx1 = 0
    idx2 = 0
    idx = 0
    for df in dfs:
        if df == field1:
            idx1 = idx
        elif df == field2:
            idx2 = idx
        idx += 1
    return idx1, idx2


def combine_arr(arr1, arr2, isAdd, month, idx):
    res = arr1
    for a2 in arr2:
        if isAdd:
            a2.insert(0, idx)
            a2.insert(0, month)
            idx += 1
        res.append(a2)
    return res


if __name__ == '__main__':
    in_excel = r"D:\university\project\excel_util\excel\东莞市万吉建材有限公司_进项明细导出文件_3月7-23日(2).xlsx"
    out_excel = r"D:\university\project\excel_util\excel\东莞市万吉建材有限公司_销货明细导出文件_3月.xlsx"
    union_excel = r"D:\university\project\excel_util\excel\待匹配_test.xlsx"
    solution(in_excel, out_excel, union_excel)

    # merged_df = pd.merge(df1, df2, how='outer', on=["货物、应税劳务及服务", "规格型号"])
    # df3 = pd.read_excel("C:\\Users\\Origami\\Desktop\\excel\\last_month.xlsx", header=1)  # 读取第二个Excel文件
    # merged_df = pd.concat([df1, df2], axis=1, ignore_index=True)  # 横向合并两个excel表
    # res_df = pd.concat([merged_df, df3], axis=0, ignore_index=True)  # 横向合并两个excel表

    # merged_df.to_excel("C:\\Users\\Origami\\Desktop\\excel\\res.xlsx", index=False)
    # print(df)  # 打印数据
    # print(out_excel.loc[1].to_dict())
