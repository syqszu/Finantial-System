import openpyxl
from openpyxl.styles import Alignment, PatternFill


def format_excel(tofill_flag, url):
    # 创建一个新的工作簿 选择活动工作表
    wb = openpyxl.load_workbook(url)
    ws = wb.active

    # 定义填充颜色
    blue_fill = PatternFill(fill_type='solid', fgColor='FFDDEBF7')
    orange_fill = PatternFill(fill_type='solid', fgColor='FFFFE699')
    green_fill = PatternFill(fill_type='solid', fgColor='B7E1CD')

    endrow = ws.max_row + 1
    # 填充A-R列为蓝色 S-AI列为橙色
    for col in range(1, 19):
        for row in range(1, endrow):
            ws.cell(row=row, column=col).fill = blue_fill
    for col in range(19, 36):
        for row in range(1, endrow):
            ws.cell(row=row, column=col).fill = orange_fill

    if tofill_flag == 0:
        for col in range(36, 39):
            for row in range(1, endrow):
                ws.cell(row=row, column=col).fill = green_fill

    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)
    ws.cell(row=1, column=1).value = "销项"
    ws.cell(row=1, column=1).fill = blue_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # 合并 S-AI列，并设置文本为"进项"
    ws.merge_cells(start_row=1, start_column=19, end_row=1, end_column=35)
    ws.cell(row=1, column=19).value = "进项"
    ws.cell(row=1, column=19).fill = orange_fill
    ws.cell(row=1, column=19).alignment = Alignment(horizontal='center')

    if tofill_flag == 0:
        ws.merge_cells(start_row=1, start_column=36, end_row=1, end_column=38)
        ws.cell(row=1, column=36).value = "进销项对比区"
        ws.cell(row=1, column=36).fill = green_fill
        ws.cell(row=1, column=36).alignment = Alignment(horizontal='center')
        ws.column_dimensions['AL'].width = 25


    bigger_columns = ['A', 'C', 'D', 'H', 'I', 'J', 'K', 'L', 'M', 'P', 'S', 'T', 'W', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD',
                      'AE','AG', 'AH']
    for column in bigger_columns:
        ws.column_dimensions[column].width = 17
    ws.column_dimensions['U'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['V'].width = 25
    ws.column_dimensions['W'].width = 25

    # 保存文件
    wb.save(url)
