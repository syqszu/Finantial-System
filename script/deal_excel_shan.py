import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment
import warnings
import os

warnings.filterwarnings('ignore')

global excel_in_url, excel_out_url, excel_tofill_url, union_outdir, union_not_outdir, union_yet_outdir, union_tofill_outdir, \
    match_out_col, match_in_col, temp_col, union_yet_df, union_not_df, union_temp_not_outdir, union_temp_yet_outdir

temp_col = pd.DataFrame()


def process_outputFilepath(outputFilepath):
    global union_outdir
    union_outdir = outputFilepath
    getOutdir()
    pre_deal()

# 传递文件名称和文件路径。
def process_excelFilepath(excelFilename, excelpath):
    global excel_in_name, excel_out_name, excel_tofill_name, excel_in_url, excel_out_url, excel_tofill_url
    if "月份进项.xlsx" in excelFilename:
        excel_in_name = excelFilename
        excel_in_url = excelpath
        print("excel_in_url", excel_in_url)
    if "月份销项.xlsx" in excelFilename:
        excel_out_name = excelFilename
        excel_out_url = excelpath
        print("excel_out_url", excel_out_url)
    if "月待填充表" in excelFilename:
        excel_tofill_name = excelFilename
        excel_tofill_url = excelpath
        print("excel_tofill_url", excel_out_url)

global process_id
def judgeprocess(filecnt):
    global process_id
    if filecnt == 2:
        process_id = 1
    elif filecnt == 3:
        process_id = 2
    print('process_id', process_id)

# 配置输出路径
def getOutdir():
    global union_not_outdir, union_yet_outdir, union_tofill_outdir, union_temp_not_outdir, union_temp_yet_outdir
    print("进项文件名字：", excel_in_name)

    month = excel_in_name[0]

    union_yet_name = month + "月匹配完成表.xlsx"
    # print(union_yet_name)
    union_not_name = month + "月匹配剩余表.xlsx"
    # print(union_not_name)
    union_tofill_name = month + "月待填充表.xlsx"
    union_temp_yet_name = month + "月临时匹配完成表.xlsx"
    union_temp_not_name = month + "月临时匹配剩余表.xlsx"

    union_not_outdir = os.path.join(union_outdir, union_not_name)
    union_yet_outdir = os.path.join(union_outdir, union_yet_name)
    union_tofill_outdir = os.path.join(union_outdir, union_tofill_name)
    union_temp_yet_outdir = os.path.join(union_outdir, union_temp_yet_name)
    union_temp_not_outdir = os.path.join(union_outdir, union_temp_not_name)

    print("union_not_outdir", union_not_outdir)
    print("union_yet_outdir", union_yet_outdir)
    print("union_tofill_outdir", union_tofill_outdir)
    print("union_temp_yet_outdir", union_temp_yet_outdir)
    print("union_temp_not_outdir", union_temp_not_outdir)


# 预处理函数 excel_in_url excel_out_url 进销项excel的绝对路径
# 1、读取表格
# 2、将表格中含有空值的行提取出来，放到temp_in里，用以后续生成“待填充表”。同时表格中删去这些含有空值的行
# 3、提取关键列，添加vis列。将税收分类编码、发票号码转化为object类型，数量转化为int类型，方便后续操作
# 4、如果有关键数值含有单引号，去除单引号。税率改为float类型的数据。增加含税单价的列。将进项表单按照含税单价排列。
# 返回只含有关键列的进销项dataframe和关键列在原excel中的列名
def pre_deal():
    global in_col, out_col, match_in_col, match_out_col
    in_col = ["税收分类编码", "发票号码", "开票日期", "销方名称", "货物、应税劳务及服务", "规格型号", "数量", "单价",
              "单位", "金额", "税率", "价税合计"]
    out_col = ["税收分类编码", "发票号码", "开票日期", "购方名称", "货物、应税劳务及服务", "规格型号", "数量", "单价",
               "单位", "金额", "税率", "价税合计", "备注"]

    # 1、读取表格
    print("开始遍历进项表格")
    df_in = pd.read_excel(excel_in_url, "Sheet1", header=1)
    match_in_col = df_in[in_col]  # pandas.core.frame.DataFrame

    # 2、如果有关键数据为空值，输出空值的行列索引并将该行额外存到一个新表中。
    # for idx, row in match_in_col.iterrows():
    #     for col in match_in_col.columns:
    #         if pd.isna(row[col]):
    #             print("表单中存在空值，请补全后重启程序。行索引-{},列索引-{}".format(idx + 3, col))
    #             temp_col.update(row)

    # 将进项表格中含有空值的行提取出来，放到temp_in里，用以后续生成“待填充表”。同时进项表格中删去这些含有空值的行
    temp_in = match_in_col[match_in_col.isna().any(axis=1)]
    temp_in["来源"] = "进项"
    match_in_col = match_in_col.dropna()
    print("进项表单遍历结束")

    # 3、提取关键列，添加vis列。将税收分类编码、发票号码转化为object类型，数量转化为int类型，方便后续操作
    match_in_col["vis"] = 0
    match_in_col["税收分类编码"] = match_in_col["税收分类编码"].astype(str)
    match_in_col["发票号码"] = match_in_col["发票号码"].astype(str)
    match_in_col["数量"] = match_in_col["数量"].astype(int)

    # 4、如果有关键数值含有单引号，去除单引号。税率改为float类型的数据。增加含税单价的列。将进项表单按照含税单价排列。
    match_in_col["税收分类编码"] = match_in_col["税收分类编码"].str.replace("'", "")
    match_in_col['税率'] = match_in_col['税率'].str.strip().str.rstrip('%').astype(float) / 100
    match_in_col["含税单价"] = match_in_col["单价"] * (1 + match_in_col["税率"])

    match_in_col = match_in_col.sort_values('含税单价')

    print("开始遍历销项表格")
    df_out = pd.read_excel(excel_out_url, "Sheet1", header=2)
    match_out_col = df_out[out_col]

    temp_out = match_out_col[match_out_col.isna().any(axis=1)]
    temp_out["来源"] = "销项"
    match_out_col = match_out_col.dropna()

    temp_col = pd.concat([temp_in, temp_out], ignore_index=True)
    print("销项表单遍历结束")
    match_out_col["vis"] = 0
    match_out_col["税收分类编码"] = match_out_col["税收分类编码"].astype(str)
    match_out_col["发票号码"] = match_out_col["发票号码"].astype(str)
    match_out_col["数量"] = match_out_col["数量"].astype(int)

    match_out_col["税收分类编码"] = match_out_col["税收分类编码"].str.replace("'", "")
    match_out_col['税率'] = match_out_col['税率'].str.strip().str.rstrip('%').astype(float) / 100
    match_out_col["含税单价"] = match_out_col["单价"] * (1 + match_in_col["税率"])

    match_out_col = match_out_col.sort_values('含税单价')

    # 如果进销项都有值，且没有上传待填充表，说明是第一次生成。
    if temp_col.empty and not excel_tofill_url:
        print("第一次生成，月待填充表为空")
        get_col(1)
    #  如果进销项都有值，但上传了待填充表，说明是第二次生成。
    if temp_col.empty and excel_tofill_url:
        print("第二次生成")
    else:
        # 进销项有的没有值，一定是第一次生成。
        print("第一次生成，月待填充表生成完毕")
        temp_col.to_excel(union_tofill_outdir, index=False)
        get_col(0)


# -*- coding: UTF-8 -*-
# 具体匹配过程，返回月匹配完成表和月匹配剩余表
# union_yet_outdir, union_not_outdir 分别是月匹配完成表和剩余表的路径。
def get_col(flag):
    # 月匹配完成表
    global union_yet_df, union_not_df
    union_yet_df = pd.DataFrame(
        columns=['开票日期', '序号', '发票号码', '税收分类编码', '货物、应税劳务及服务', '规格型号', '单位',
                 '上月原数量', '上月程序出库数', '上月人工出库数', '本月剩余数量', '不含税单价',
                 '不含税金额', '含税单价', '含税总金额', '供应商',  # 16 销项

                 '开票日期1', '序号1', '发票号码1', '税收分类编码1', '货物、应税劳务及服务1', '规格型号1', '单位1',
                 '上月原数量1', '上月程序出库数1', '上月人工出库数1', '本月剩余数量1', '不含税单价1',
                 '不含税金额1', '含税单价1', '含税总金额1', '销方名称'  # 进项
                 ])
    union_not_df = union_yet_df.copy()  # 月匹配剩余表
    union_temp_not_df = union_yet_df.copy()  # 月临时匹配剩余表
    union_temp_yet_df = union_yet_df.copy()  # 月临时匹配完成表
    id = 1
    if flag == 1:
        # 暴力循环
        print("月匹配完成表开始生成")
    else:
        print("月临时匹配完成表开始生成")
    for idx_out, row_out in match_out_col.iterrows():
        for idx_in, row_in in match_in_col.iterrows():

            date_out = row_out["开票日期"]  # 2023-03-05
            date_in = row_in["开票日期"]
            number_out = row_out["发票号码"]
            number_in = row_in["发票号码"]
            loads_out = row_out["货物、应税劳务及服务"]
            loads_in = row_in["货物、应税劳务及服务"]
            size_out = row_out["规格型号"]
            size_in = row_in["规格型号"]
            unit_out = row_out["单位"]
            unit_in = row_in["单位"]
            origin_nums_out = row_out["数量"]  # 销项上月原数量
            origin_nums_in = row_in["数量"]  # 进项上月原数量
            notax_perprice_out = row_out["单价"]   # 销项不含税单价
            notax_perprice_in = row_in["单价"]  # 进项不含税单价

            notax_money_out = row_out["金额"]  # 销项不含税金额
            notax_money_in = row_in["金额"]  # 进项不含税金额

            tax_out = row_out["税率"]  # 销项税率
            tax_in = row_in["税率"]  # 进项税率

            # print(type(row_in["税率"]))
            # print(row_in["税率"])

            tax_perprice_out = row_out["含税单价"]  # 销项含税单价
            tax_perprice_in = row_in["含税单价"]  # 进项含税单价

            tax_money_out = row_out["价税合计"]  # 销项含税总金额
            tax_money_in = row_in["价税合计"]  # 进项含税总金额

            name_in = row_in["销方名称"]  # 销方名称
            name_out = row_out["购方名称"]  # 供应商

            difference = abs(tax_perprice_in - tax_perprice_out)
            percentage_differ = (difference / tax_perprice_in) * 100

            case1 = row_out[0] == row_in[0] and row_out["vis"] == 0 and row_in["vis"] == 0

            # front7_in = "{:.7f}".format(row_in[0])
            front7_in = row_in[0][0:6]
            # print(front7_in)
            # print(type(front7_in))

            # front7_out = "{:.7f}".format(row_out[0])
            front7_out = row_out[0][0:6]

            case2 = front7_in == front7_out and row_out["vis"] == 0 and row_in["vis"] == 0 and percentage_differ <= 10

            # 第一次遍历，将税收分类编码完美匹配的vis打上1的标记   前七位相同，含税单价相差不超过10%的vis打上2的标记
            if case1 or case2:

                yet_row = {"开票日期": date_out, "序号": id, "发票号码": number_out,
                           "税收分类编码": row_out[0], "货物、应税劳务及服务": loads_out, "规格型号": size_out,
                           "单位": unit_out, "上月原数量": origin_nums_out}

                # value = [date_out, number_out, loads_out, size_out, unit_out, origin_nums_out]
                # union_yet_df = pd.DataFrame(np.insert(union_yet_df.values, id-1, values=value), axis=0)

                outbound_pro_nums_out = outbound_pro_nums_in = 0  # 上月程序出库数
                remain_nums_out = remain_nums_in = 0  # 本月剩余数量

                if origin_nums_in < origin_nums_out:  # 进项上月原数量<销项上月原数量
                    outbound_pro_nums_out = outbound_pro_nums_in = origin_nums_in  # 销进项上月程序出库数=进项上月原数量
                    remain_nums_out = origin_nums_out - origin_nums_in  # 销项本月剩余量=销项上月原数量-进项上月原数量
                    remain_nums_in = 0  # 进项本月剩余量=0
                    temp_row = {}
                    for col in out_col:
                        temp_row.update({col: row_out[col]})
                    temp_row.update({"数量": remain_nums_out})
                    match_out_col.loc[len(match_out_col)] = temp_row

                else:
                    outbound_pro_nums_out = outbound_pro_nums_in = origin_nums_out
                    remain_nums_in = origin_nums_in - origin_nums_out
                    remain_nums_out = 0
                    temp_row = {}
                    for col in in_col:
                        temp_row.update({col: row_in[col]})
                    temp_row.update({"数量": remain_nums_in})
                    match_in_col.loc[len(match_in_col)] = temp_row

                yet_row.update(
                    {"上月程序出库数": outbound_pro_nums_out, "上月人工出库数": 0, "本月剩余数量": remain_nums_out})

                yet_row.update(
                    {"不含税单价": notax_perprice_out, "不含税金额": notax_money_out, "含税单价": tax_perprice_out,
                     "含税总金额": tax_money_out, "供应商": name_out})
                yet_row.update({"开票日期1": date_in, "序号1": id, "发票号码1": number_in, "税收分类编码1": row_in[0],
                                "货物、应税劳务及服务1": loads_in, "规格型号1": size_in, "单位1": unit_in,
                                "上月原数量1": origin_nums_in, "上月程序出库数1": outbound_pro_nums_in,
                                "上月人工出库数1": 0,
                                "本月剩余数量1": remain_nums_in, "不含税单价1": notax_perprice_in,
                                "不含税金额1": notax_money_in,
                                "含税单价1": tax_perprice_in, "含税总金额1": tax_money_in, "销方名称": name_in})

                union_yet_df = union_yet_df.append(yet_row, ignore_index=True)

                if case1:
                    match_out_col.at[idx_out, "vis"] = match_in_col.at[idx_in, "vis"] = 1
                if case2:
                    match_out_col.at[idx_out, "vis"] = match_in_col.at[idx_in, "vis"] = 2

                id += 1
    if flag == 1:
        union_yet_df.to_excel(union_yet_outdir, index=False)
        create_union_excel(union_yet_outdir)
        print("月匹配完成表生成完毕")
        print("月匹配剩余表开始生成")
    else:
        union_yet_df.to_excel(union_temp_yet_outdir, index=False)
        create_union_excel(union_temp_yet_outdir)
        print("月临时匹配完成表生成完毕")
        print("月临时匹配剩余表开始生成")
    # 生成未匹配表
    id = 1
    for idx_out, row_out in match_out_col.iterrows():
        if row_out["vis"] == 0:
            date_out = row_out["开票日期"]  # 2023-03-05
            number_out = row_out["发票号码"]
            loads_out = row_out["货物、应税劳务及服务"]
            size_out = row_out["规格型号"]
            unit_out = row_out["单位"]
            origin_nums_out = remain_nums_out = row_out["数量"]  # 销项上月原数量=本月剩余数量=销项文件的数量
            notax_perprice_out = row_out["单价"]  # 销项不含税单价
            notax_money_out = row_out["金额"]  # 销项不含税金额
            tax_perprice_out = row_out["含税单价"]  # 销项含税单价
            tax_money_out = row_out["价税合计"]  # 销项含税总金额
            name_out = row_out["购方名称"]  # 供应商
            not_row = {"开票日期": date_out, "序号": id, "发票号码": number_out,
                       "税收分类编码": row_out[0], "货物、应税劳务及服务": loads_out, "规格型号": size_out,
                       "单位": unit_out, "上月原数量": origin_nums_out, "上月程序出库数": 0,
                       "上月人工出库数": 0, "本月剩余数量": remain_nums_out, "不含税单价": notax_perprice_out,
                       "不含税金额": notax_money_out, "含税单价": tax_perprice_out,
                       "含税总金额": tax_money_out, "供应商": name_out}
            union_not_df = union_not_df.append(not_row, ignore_index=True)
            id += 1
        # union_not_df = union_not_df.append(not_row, ignore_index=True)

    id = 1
    for idx_in, row_in in match_in_col.iterrows():
        if row_in["vis"] == 0:
            date_in = row_in["开票日期"]
            number_in = row_in["发票号码"]
            loads_in = row_in["货物、应税劳务及服务"]
            size_in = row_in["规格型号"]
            unit_in = row_in["单位"]
            origin_nums_in = remain_nums_in = row_in["数量"]  # 进项上月原数量=本月剩余数量=销项文件的数量
            notax_perprice_in = row_in["单价"]  # 进项不含税单价
            tax_perprice_in = row_in["含税单价"]  # 进项含税单价
            notax_money_in = row_in["金额"]  # 进项不含税金额
            tax_money_in = row_in["价税合计"]  # 进项含税总金额
            name_in = row_in["销方名称"]  # 销方名称
            not_row = {"开票日期1": date_in, "序号1": id, "发票号码1": number_in, "税收分类编码1": row_in[0],
                       "货物、应税劳务及服务1": loads_in, "规格型号1": size_in, "单位1": unit_in,
                       "上月原数量1": origin_nums_in, "上月程序出库数1": 0,
                       "上月人工出库数1": 0,
                       "本月剩余数量1": remain_nums_in, "不含税单价1": notax_perprice_in,
                       "不含税金额1": notax_money_in,
                       "含税单价1": tax_perprice_in, "含税总金额1": tax_money_in, "销方名称": name_in}
            union_not_df = union_not_df.append(not_row, ignore_index=True)
    # print(union_yet_df)
    # 1080207990000000000 是完全一致的
    if flag == 1:
        union_not_df.to_excel(union_not_outdir, index=False)
        create_union_excel(union_not_outdir)
        print("月匹配剩余表生成完毕")
    else:
        union_not_df.to_excel(union_temp_not_outdir, index=False)
        create_union_excel(union_temp_not_outdir)
        print("月临时匹配剩余表生成完毕")


def create_union_excel(url):
    # 创建一个新的工作簿 选择活动工作表
    wb = openpyxl.load_workbook(url)
    ws = wb.active

    # 定义填充颜色
    blue_fill = PatternFill(fill_type='solid', fgColor='FFDDEBF7')
    orange_fill = PatternFill(fill_type='solid', fgColor='FFFFE699')

    endrow = ws.max_row + 1
    # 填充A-P列为蓝色 Q-AF列为橙色
    for col in range(1, 17):
        for row in range(1, endrow):
            ws.cell(row=row, column=col).fill = blue_fill
    for col in range(17, 33):
        for row in range(1, endrow):
            ws.cell(row=row, column=col).fill = orange_fill

    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    ws.cell(row=1, column=1).value = "销项"
    ws.cell(row=1, column=1).fill = blue_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # 合并Q-AF列，并设置文本为"进项"
    ws.merge_cells(start_row=1, start_column=17, end_row=1, end_column=32)
    ws.cell(row=1, column=17).value = "进项"
    ws.cell(row=1, column=17).fill = orange_fill
    ws.cell(row=1, column=17).alignment = Alignment(horizontal='center')

    # 定义列名列表
    # column_names = ["开票日期", "序号", "发票号码", "税收分类编码", "货物、应税劳务及服务", "规格型号", "单位",
    #                 "上月原数量", "上月程序出库数",
    #                 "上月人工出库数", "本月剩余数量", "不含税单价", "不含税金额", "含税单价", "含税总金额", "销方名称",
    #                 "开票日期1", "序号1", "发票号码1", "税收分类编码1", "货物、应税劳务及服务1", "规格型号1", "单位1",
    #                 "上月原数量1", "上月程序出库数1",
    #                 "上月人工出库数1", "本月剩余数量1", "不含税单价1", "不含税金额1", "含税单价1", "含税总金额1",
    #                 "供应商1"]
    # 设置第二行的列名
    # for index, name in enumerate(column_names):
    #     ws.cell(row=3, column=index + 1).value = name

    bigger_columns = ['A', 'D', 'H', 'I', 'J', 'K', 'L', 'M', 'O', 'Q', 'T', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD',
                      'AE']
    for column in bigger_columns:
        ws.column_dimensions[column].width = 17
    ws.column_dimensions['U'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['AF'].width = 25

    # 保存文件
    wb.save(url)

# if __name__ == '__main__':
#     print("开始配置输出路径")
#     getOutdir()
#     print("得到输出路径")
# print("开始预处理")
# pre_deal()
# print("预处理完毕")
# get_col()
# getOutdir()
# create_union_excel()
